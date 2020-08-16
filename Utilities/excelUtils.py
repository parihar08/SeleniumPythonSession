import openpyxl


def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)


def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,colNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=colNum).value

def writeData(file,sheetName,rownum,colNum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    data = sheet.cell(row=rownum, column=colNum).value
    workbook.save(file)